## Import Package "Openpyxl" (ÉÉRST VIA PIP DOWNLOADEN EN INSTALLEREN)
import openpyxl

## Load file
File = openpyxl.load_workbook("C:/Users/Klaar/OneDrive/Bureaublad/TestDataRead.xlsx")

## Geef alle bestaande tabblad-namen in deze file.
# print(File.sheetnames)

## Geef de actieve tabblad-naam.
# print("Actieve tabblad is "+File.active.title)

## Ga naar het tabblad waar je in wilt gaan werken.
FileTab = File['Blad1']
# print(FileTab.title)

## Manier 1 om de inhoud van een cell op te vragen.
print(FileTab['A3'].value)

## Manier 2 om de inhoud van een cell op te vragen.
Cell_A3= FileTab.cell(3,1)
print(Cell_A3.value)

## Manier 3 om de inhoud van een cell op te vragen.
Cell_A3= FileTab.cell(column=1, row=3)
print(Cell_A3.value)

## Manier 4 om de inhoud van een cell op te vragen.
Cell_A3= FileTab.cell(row=3, column=1)
print(Cell_A3.value)

## Vraag de rij-nummer en  kolom-nummer op van "Cell_A3".
print(Cell_A3.row)
print(Cell_A3.column)

## Vraag het aantal rijen en kkolommen op van tabblad "Blad1".
AantalRijen = FileTab.max_row
AantalKolommen = FileTab.max_column

print("Totaal aantal rijen in tabblad " + str(FileTab) + " is: " + str(AantalRijen))
print("Totaal aantal kolommen in tabblad " + str(FileTab) + " is: " + str(AantalKolommen))

# Manier 1: Vraag nu alle data in tabblad "Blad1" op.
for Rijen in range(1,AantalRijen+1):
    for Kolommen in range(1,AantalKolommen+1):
        Aantallen=FileTab.cell(Rijen,Kolommen)
        print(Aantallen.value)

# Manier 2: Vraag nu alle data in tabblad "Blad1" op.
# for Tabblad in FileTab['A1':'C4']:
#     for Cellen in Tabblad:
#        print(Cellen.value)

