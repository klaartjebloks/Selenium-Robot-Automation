def CreateExcelFile(Count_Aantal_Vrienden, Count_Aantal_Active_Vrienden):
    ## Import Package "Openpyxl" (ÉÉRST VIA PIP DOWNLOADEN EN INSTALLEREN)
    import openpyxl

    ## Ken een pad en NIEUWE file naam toe aan variabele "FileName".
    FileName = "C:/Users/Klaar/OneDrive/Bureaublad/Facebook_Scripts/FB_Vrienden.xlsx"

    ## Creëeer een nieuw Workbook. Een Workbook wordt later opgeslagen als een ExcelFile.
    wb = openpyxl.Workbook()

    ## Ga na actieve tabblad van dit Workbook.
    Tabblad = wb.active

    ## Hernoem de default titel van het active tabblad naar onderstaande titel.
    Tabblad.title="Vrienden (" + str(Count_Aantal_Vrienden) + "-" + str(Count_Aantal_Active_Vrienden) + ")"

    ## Schrijf onderstaande waarde in de aangegeven cellen van dit tabblad weg.
    Tabblad["A1"].value = "ProfielURL"
    Tabblad["B1"].value = "Naam"
    Tabblad["C1"].value = "Geslacht"
    Tabblad["D1"].value = "Geboortedatum"
    Tabblad["E1"].value = "Huidige VriendStatus"
    Tabblad["F1"].value = "Nieuwe VriendStatus"
    Tabblad["G1"].value = "Ontvrienden"

    ## Sla het aangemaakte workbook op als deze file.
    wb.save(FileName)
    wb.close

def AddFriend(Profiel_URL, Naam, RegelNr):
    ## Import Package "Openpyxl" (ÉÉRST VIA PIP DOWNLOADEN EN INSTALLEREN)
    import openpyxl

    ## Ken een pad en NIEUWE file naam toe aan variabele "FileName".
    FileName = "C:/Users/Klaar/OneDrive/Bureaublad/Facebook_Scripts/FB_Vrienden.xlsx"

    ## Open het BESTAANDE Workbook.
    wb = openpyxl.load_workbook(FileName)

    ## Ga na actieve tabblad van dit Workbook.
    Tabblad = wb.active

    ## Schrijf onderstaande waarde in de aangegeven cellen van dit tabblad weg.
    Tabblad["A" + str(RegelNr) + ""].value = Profiel_URL
    Tabblad["B" + str(RegelNr) + ""].value = Naam

    wb.save(FileName)
    wb.close

    ## Geef het profiel URL terug.
    return Profiel_URL

def FillFriend(Geslacht, Geboorte_Datum, Vriend_Type, RegelNr):
    ## Import Package "Openpyxl" (ÉÉRST VIA PIP DOWNLOADEN EN INSTALLEREN)
    import openpyxl

    ## Ken een pad en NIEUWE file naam toe aan variabele "FileName".
    FileName = "C:/Users/Klaar/OneDrive/Bureaublad/Facebook_Scripts/FB_Vrienden.xlsx"

    ## Open het BESTAANDE Workbook.
    wb = openpyxl.load_workbook(FileName)

    ## Ga na actieve tabblad van dit Workbook.
    Tabblad = wb.active

    ## Schrijf onderstaande waarde in de aangegeven cellen van dit tabblad weg.
    Tabblad["C" + str(RegelNr) + ""].value = Geslacht
    Tabblad["D" + str(RegelNr) + ""].value = Geboorte_Datum
    Tabblad["E" + str(RegelNr) + ""].value = Vriend_Type

    wb.save(FileName)
    wb.close

    ## ....
    return Geboorte_Datum

def CountRows():
    ## Import Package "Openpyxl" (ÉÉRST VIA PIP DOWNLOADEN EN INSTALLEREN)
    import openpyxl

    ## Ken een pad en NIEUWE file naam toe aan variabele "FileName".
    FileName = "C:/Users/Klaar/OneDrive/Bureaublad/Facebook_Scripts/FB_Vrienden.xlsx"

    ## Open het BESTAANDE Workbook.
    wb = openpyxl.load_workbook(FileName)

    ## Ga na actieve tabblad van dit Workbook.
    Tabblad = wb.active

    ## Vraag het aantal rijen op van tabblad.
    AantalRijen = Tabblad.max_row
    return AantalRijen

    wb.save(FileName)
    wb.close

def SplitProfielURL(URL_String):
    ## Splits op het "?"
    Profiel_url = URL_String.split("?")
    ## Geef de string vóór het "?" terug.
    return Profiel_url[0]

def RijRegelOphogen(j):
    Rij_Nr = int(j) + 1
    return Rij_Nr

def GetURL(RegelNr):
    ## Import Package "Openpyxl" (ÉÉRST VIA PIP DOWNLOADEN EN INSTALLEREN)
    import openpyxl

    ## Ken een pad en NIEUWE file naam toe aan variabele "FileName".
    FileName = "C:/Users/Klaar/OneDrive/Bureaublad/Facebook_Scripts/FB_Vrienden.xlsx"

    ## Open het BESTAANDE Workbook.
    wb = openpyxl.load_workbook(FileName)

    ## Ga na actieve tabblad van dit Workbook.
    Tabblad = wb.active

    ## Vraag nu de data (= vriend profiel URL) in kolom A op.
    return(Tabblad["A"+ str(RegelNr) +""].value)

def GetHuidigeVriendStatus(RegelNr):
    ## Import Package "Openpyxl" (ÉÉRST VIA PIP DOWNLOADEN EN INSTALLEREN)
    import openpyxl

    ## Ken een pad en NIEUWE file naam toe aan variabele "FileName".
    FileName = "C:/Users/Klaar/OneDrive/Bureaublad/Facebook_Scripts/FB_Vrienden.xlsx"

    ## Open het BESTAANDE Workbook.
    wb = openpyxl.load_workbook(FileName)

    ## Ga na actieve tabblad van dit Workbook.
    Tabblad = wb.active

    ## Vraag nu de data (= vriend profiel URL) in kolom A op.
    return(Tabblad["E"+ str(RegelNr) +""].value)

def GetNieuweVriendStatus(RegelNr):
    ## Import Package "Openpyxl" (ÉÉRST VIA PIP DOWNLOADEN EN INSTALLEREN)
    import openpyxl

    ## Ken een pad en NIEUWE file naam toe aan variabele "FileName".
    FileName = "C:/Users/Klaar/OneDrive/Bureaublad/Facebook_Scripts/FB_Vrienden.xlsx"

    ## Open het BESTAANDE Workbook.
    wb = openpyxl.load_workbook(FileName)

    ## Ga na actieve tabblad van dit Workbook.
    Tabblad = wb.active

    ## Vraag nu de data (= vriend profiel URL) in kolom A op.
    return(Tabblad["F"+ str(RegelNr) +""].value)

def GetVriendNaam(RegelNr):
    ## Import Package "Openpyxl" (ÉÉRST VIA PIP DOWNLOADEN EN INSTALLEREN)
    import openpyxl

    ## Ken een pad en NIEUWE file naam toe aan variabele "FileName".
    FileName = "C:/Users/Klaar/OneDrive/Bureaublad/Facebook_Scripts/FB_Vrienden.xlsx"

    ## Open het BESTAANDE Workbook.
    wb = openpyxl.load_workbook(FileName)

    ## Ga na actieve tabblad van dit Workbook.
    Tabblad = wb.active

    ## Vraag nu de data (= vriend profiel URL) in kolom A op.
    return(Tabblad["B"+ str(RegelNr) +""].value)

def Ontvrienden(RegelNr):
    ## Import Package "Openpyxl" (ÉÉRST VIA PIP DOWNLOADEN EN INSTALLEREN)
    import openpyxl

    ## Ken een pad en NIEUWE file naam toe aan variabele "FileName".
    FileName = "C:/Users/Klaar/OneDrive/Bureaublad/Facebook_Scripts/FB_Vrienden.xlsx"

    ## Open het BESTAANDE Workbook.
    wb = openpyxl.load_workbook(FileName)

    ## Ga na actieve tabblad van dit Workbook.
    Tabblad = wb.active

    ## Vraag nu de data (= vriend profiel URL) in kolom A op.
    return(Tabblad["G"+ str(RegelNr) +""].value)

def CountChangedRows(Vriend_Status, Type_Vriend):
    ## Import Package "Openpyxl" (ÉÉRST VIA PIP DOWNLOADEN EN INSTALLEREN)
    import openpyxl

    ## Ken een pad en NIEUWE file naam toe aan variabele "FileName".
    FileName = "C:/Users/Klaar/OneDrive/Bureaublad/Facebook_Scripts/FB_Vrienden.xlsx"

    ## Open het BESTAANDE Workbook.
    wb = openpyxl.load_workbook(FileName)

    ## Ga na actieve tabblad van dit Workbook.
    Tabblad = wb.active

    ## Vraag het aantal rijen op van tabblad.
    if Vriend_Status != Type_Vriend:
        AantalGewijzigdeRijen = Tabblad.max_row

    return AantalGewijzigdeRijen

    wb.save(FileName)
    wb.close